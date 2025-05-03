from openpyxl import load_workbook
from os.path import abspath


class XlsxOperator(object):
    def __init__(self, file_name: str, data: list, horizontal: bool = False):
        self.file_name = abspath(file_name)
        self.workbook = load_workbook(file_name)
        self.sheet = self.workbook.active

        # Find the last column in the sheet
        self.last_column = self.sheet.max_column
        if self.last_column == 1 and not horizontal:
            self.create_headings(data, horizontal)
            self.last_column += 1
        # Determine the next column to write data (last_column + 1)
        self.next_column = self.last_column + 1

        # Find the last row in the sheet
        self.last_row = self.sheet.max_row
        if self.last_row == 1 and horizontal:
            self.create_headings(data, horizontal)
            self.last_row += 1
        # Determine the next row to write data (last_row + 1)
        self.next_row = self.last_row + 1

    def create_headings(self, data: list, horizontal: bool = False):
        """
        Create headings when the sheet is empty.

        :param data: Data that is needed to write to the sheet
        :param horizontal: If it needs to write the data horizontally. Defaults to False.
        """

        if not horizontal:
            for row, value in enumerate(data, start=1):
                self.sheet.cell(row=row, column=self.last_column, value=value)
        else:
            for column, value in enumerate(data, start=1):
                self.sheet.cell(row=self.last_row, column=column, value=value)

    def add_data(self, data: list, horizontal: bool = False):
        """
        Add the data to the sheet.

        :param data: Data that is needed to write to the sheet
        :param horizontal: If it needs to write the data horizontally. Defaults to False.
        """

        if not horizontal:
            for row, value in enumerate(data, start=1):
                self.sheet.cell(row=row, column=self.next_column, value=value)
                self.next_column += 1
        else:
            for column, value in enumerate(data, start=1):
                self.sheet.cell(row=self.next_row, column=column, value=value)
                self.next_row += 1

    def save_sheet(self):
        # Save the workbook
        self.workbook.save(self.file_name)
